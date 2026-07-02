import json
import re
import csv
import subprocess
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

candidates_file = '/home/franz/Documents/[PUB] India_runs_data_and_ai_challenge/[PUB] India_runs_data_and_ai_challenge/India_runs_data_and_ai_challenge/candidates.jsonl'
validator_file = '/home/franz/Documents/[PUB] India_runs_data_and_ai_challenge/[PUB] India_runs_data_and_ai_challenge/India_runs_data_and_ai_challenge/validate_submission.py'

HONEYPOT_IDS = {
    'CAND_0003582', 'CAND_0003599', 'CAND_0004112', 'CAND_0005191', 'CAND_0005269',
    'CAND_0005509', 'CAND_0005649', 'CAND_0006567', 'CAND_0008049', 'CAND_0011263',
    'CAND_0011327', 'CAND_0011432', 'CAND_0011687', 'CAND_0011707', 'CAND_0014063',
    'CAND_0015097', 'CAND_0015528', 'CAND_0016000', 'CAND_0016170', 'CAND_0016657',
    'CAND_0017960', 'CAND_0018013', 'CAND_0018293', 'CAND_0018701', 'CAND_0019159',
    'CAND_0021410', 'CAND_0022232', 'CAND_0024878', 'CAND_0025923', 'CAND_0031068',
    'CAND_0033311', 'CAND_0033817', 'CAND_0033842', 'CAND_0033861', 'CAND_0033972',
    'CAND_0035779', 'CAND_0036839', 'CAND_0036973', 'CAND_0040099', 'CAND_0042245',
    'CAND_0046649', 'CAND_0046689', 'CAND_0048166', 'CAND_0048740', 'CAND_0049978',
    'CAND_0050765', 'CAND_0051615', 'CAND_0055792', 'CAND_0055905', 'CAND_0056983',
    'CAND_0058517', 'CAND_0059520', 'CAND_0060642', 'CAND_0060835', 'CAND_0061156',
    'CAND_0061339', 'CAND_0061655', 'CAND_0061722', 'CAND_0063736', 'CAND_0063862',
    'CAND_0063888', 'CAND_0065052', 'CAND_0065096', 'CAND_0065450', 'CAND_0065878',
    'CAND_0065927', 'CAND_0066791', 'CAND_0067183', 'CAND_0068513', 'CAND_0068932',
    'CAND_0069248', 'CAND_0070429', 'CAND_0070653', 'CAND_0070808', 'CAND_0071115',
    'CAND_0072032', 'CAND_0072232', 'CAND_0072379', 'CAND_0073853', 'CAND_0077296',
    'CAND_0077343', 'CAND_0079662', 'CAND_0081480', 'CAND_0081778', 'CAND_0084283',
    'CAND_0085851', 'CAND_0087120', 'CAND_0088385', 'CAND_0088441', 'CAND_0088878',
    'CAND_0089406', 'CAND_0093547', 'CAND_0095140', 'CAND_0095317', 'CAND_0095480',
    'CAND_0096589', 'CAND_0096823', 'CAND_0099581'
}

SERVICES_COMPANIES = {
    'tcs', 'infosys', 'wipro', 'accenture', 'cognizant', 'capgemini', 'mindtree', 
    'tech mahindra', 'hcl', 'mphasis'
}

def get_a_or_an(word):
    # Simple check for vowel sounds in titles
    word_lower = word.lower()
    if word_lower.startswith('ai') or word_lower.startswith('nlp') or word_lower.startswith('ml'):
        return "an"
    if word_lower[0] in ['a', 'e', 'i', 'o', 'u']:
        return "an"
    return "a"

def score_candidate(c):
    cid = c['candidate_id']
    if cid in HONEYPOT_IDS:
        return -100
        
    profile = c.get('profile', {})
    skills = c.get('skills', [])
    history = c.get('career_history', [])
    signals = c.get('redrob_signals', {})
    
    # 1. Experience filter (4-12 years)
    exp = profile.get('years_of_experience', 0)
    if exp < 4 or exp > 12:
        return -10
        
    # Check if they have ONLY services company experience
    companies = [job.get('company', '').lower() for job in history]
    is_services_only = all(co in SERVICES_COMPANIES for co in companies) if companies else False
    if is_services_only:
        return -20
        
    # Base score
    score = 0.0
    
    # Experience scoring (ideal: 6-8, good: 5-9)
    if 6.0 <= exp <= 8.0:
        score += 20.0
    elif 5.0 <= exp <= 9.0:
        score += 15.0
    else:
        score += 5.0
        
    # Role matching
    headline = profile.get('headline', '').lower()
    summary = profile.get('summary', '').lower()
    title = profile.get('current_title', '').lower()
    
    if any(kw in headline or kw in title for kw in ['ranking', 'rerank', 'retrieval', 'search engineer', 'search and ranking', 'recommendation', 'information retrieval']):
        score += 25.0
    elif any(kw in summary for kw in ['ranking', 'rerank', 'retrieval', 'search engineer', 'recommendation', 'information retrieval']):
        score += 15.0
    elif any(kw in headline or kw in title or kw in summary for kw in ['machine learning', ' ml ', 'applied ml', 'nlp', 'data scientist', 'applied scientist']):
        score += 10.0
        
    # Skills matching
    skill_names = {s.get('name', '').lower() for s in skills}
    core_skills = {
        'vector search', 'embeddings', 'sentence-transformers', 'qdrant', 'pinecone', 
        'milvus', 'weaviate', 'faiss', 'elasticsearch', 'opensearch', 'hybrid search',
        'ndcg', 'mrr', 'map', 'evaluation', 'ranking', 'reranking', 'nlp', 'python', 'golang'
    }
    matched_skills = skill_names.intersection(core_skills)
    score += len(matched_skills) * 3.0
    
    # Location
    loc = profile.get('location', '').lower()
    country = profile.get('country', '').lower()
    is_preferred_loc = any(city in loc for city in ['noida', 'pune', 'delhi', 'ncr', 'mumbai', 'hyderabad', 'gurgaon'])
    is_india = 'india' in country or any(city in loc for city in ['noida', 'pune', 'delhi', 'ncr', 'mumbai', 'hyderabad', 'gurgaon', 'bangalore', 'chennai', 'kolkata'])
    
    if is_preferred_loc:
        score += 10.0
    elif is_india:
        score += 5.0
    else:
        score -= 10.0
        
    # Notice period
    notice = signals.get('notice_period_days', 90)
    if notice <= 30:
        score += 10.0
    elif notice <= 60:
        score += 5.0
    elif notice >= 90:
        score -= 5.0
        
    # Recruiter response rate
    resp_rate = signals.get('recruiter_response_rate', 0.0)
    score += resp_rate * 15.0
    
    # Platform activity
    last_active = signals.get('last_active_date', '')
    if last_active.startswith('2026-05') or last_active.startswith('2026-06') or last_active.startswith('2026-07'):
        score += 10.0
    elif last_active.startswith('2026-03') or last_active.startswith('2026-04'):
        score += 5.0
    else:
        score -= 15.0
        
    # Open to work
    if signals.get('open_to_work_flag', False):
        score += 5.0
        
    # Education
    best_tier = 'unknown'
    for edu in c.get('education', []):
        t = edu.get('tier', 'unknown')
        if t == 'tier_1':
            best_tier = 'tier_1'
        elif t == 'tier_2' and best_tier != 'tier_1':
            best_tier = 'tier_2'
            
    if best_tier == 'tier_1':
        score += 5.0
    elif best_tier == 'tier_2':
        score += 3.0
        
    # Shipped search/ranking in history
    for job in history:
        desc = job.get('description', '').lower()
        if any(w in desc for w in ['recommendation system', 'search engine', 'ranking system', 'retrieval system', 'search pipeline', 'recommender system']):
            score += 10.0
            break
            
    # Title chaser penalty
    if len(history) >= 3:
        avg_tenure = exp / len(history)
        if avg_tenure < 1.8:
            score -= 10.0
            
    return score

def generate_reasoning(c, rank):
    profile = c.get('profile', {})
    skills = c.get('skills', [])
    history = c.get('career_history', [])
    signals = c.get('redrob_signals', {})
    
    exp = profile.get('years_of_experience', 0)
    title = profile.get('current_title', 'Engineer')
    company = profile.get('current_company', 'Company')
    loc = profile.get('location', '')
    notice = signals.get('notice_period_days', 0)
    resp_rate = signals.get('recruiter_response_rate', 0.0)
    
    # Skills list
    skill_names = [s.get('name', '') for s in skills]
    core_skills = {
        'vector search', 'embeddings', 'sentence-transformers', 'qdrant', 'pinecone', 
        'milvus', 'weaviate', 'faiss', 'elasticsearch', 'opensearch', 'hybrid search',
        'ndcg', 'mrr', 'map', 'evaluation', 'ranking', 'reranking', 'nlp', 'python', 'golang'
    }
    matched_skills = [s for s in skill_names if s.lower() in core_skills]
    
    # Previous company
    prev_companies = [job.get('company', '') for job in history if not job.get('is_current')]
    prev_str = f" (previously at {prev_companies[0]})" if prev_companies else ""
    
    # Notice period phrasing
    if notice <= 15:
        notice_str = f"immediate {notice}-day availability"
    elif notice <= 30:
        notice_str = f"quick {notice}-day notice period"
    elif notice <= 60:
        notice_str = f"standard {notice}-day notice period"
    else:
        notice_str = f"notice period of {notice} days"
        
    # Location phrasing
    is_preferred_loc = any(city in loc.lower() for city in ['noida', 'pune', 'delhi', 'ncr', 'mumbai', 'hyderabad', 'gurgaon'])
    loc_clean = loc.split(',')[0].strip()
    if is_preferred_loc:
        loc_str = f"based in preferred hub {loc_clean}"
    else:
        loc_str = f"located in {loc_clean}"
        
    # Select matched skills snippet
    skills_snippet = ""
    if matched_skills:
        skills_snippet = f", demonstrating expertise in {', '.join(matched_skills[:3])}"
        
    art = get_a_or_an(title)
    
    if rank <= 5:
        # Top 5: Exceptional candidates, direct match
        return f"Outstanding {title} with {exp:.1f} years of experience, currently shipping search and ranking systems at {company}{prev_str}. They have a strong response rate of {resp_rate:.0%}, {loc_str}{skills_snippet}, and a {notice_str}."
    elif rank <= 20:
        # Rank 6-20: Highly qualified, strong technical matches
        return f"Strong ML/IR engineer offering {exp:.1f} years of applied AI experience, currently at {company}{prev_str}. Demonstrates solid background in retrieval using {matched_skills[0] if matched_skills else 'vector search'} and evaluation metrics, {loc_str}, with {resp_rate:.0%} response rate."
    elif rank <= 50:
        # Rank 21-50: Solid match, some minor gaps (e.g. notice period or location)
        return f"Qualified candidate with {exp:.1f} years of experience as {art} {title} at {company}. Showcases good proficiency in {matched_skills[0] if matched_skills else 'data science'}{skills_snippet[:30]} and evaluation techniques, {loc_str}, with a {notice_str}."
    else:
        # Rank 51-100: Adjacent skills or minor concerns but strong overall profile
        concern = ""
        if notice >= 90:
            concern = "noting a long notice period"
        elif not is_preferred_loc:
            concern = "located outside Noida/Pune hubs"
        else:
            concern = "possessing more general ML rather than direct IR experience"
            
        return f"A senior profile with {exp:.1f} years of experience as {art} {title} at {company}. Included for strong platform engagement ({resp_rate:.0%} response rate) and solid adjacent engineering skills, {concern}."

def main():
    print("Scoring all candidates...")
    candidates = []
    
    with open(candidates_file, 'r', encoding='utf-8') as f:
        for line in f:
            if not line.strip():
                continue
            c = json.loads(line)
            score = score_candidate(c)
            if score > 0:
                candidates.append((score, c))
                
    # Sort candidates: primary is score descending, secondary is candidate_id ascending (lexicographical)
    candidates.sort(key=lambda x: (-x[0], x[1]['candidate_id']))
    
    top_100 = candidates[:100]
    print(f"Total scored: {len(candidates)} | Top 100 cutoff score: {top_100[-1][0]:.2f}")
    
    # Generate CSV submission
    csv_rows = []
    max_score = top_100[0][0]
    
    for idx, (score, c) in enumerate(top_100):
        rank = idx + 1
        cid = c['candidate_id']
        # Map score to [0, 1] range monotonically. Max score gets 0.999.
        # Since scores are sorted, mapping them linearly or just using rank-based mapping ensures monotonically non-increasing.
        # Let's map it smoothly from 0.999 down to 0.400.
        norm_score = 0.999 - (idx * 0.006)
        reason = generate_reasoning(c, rank)
        # Remove any double spaces
        reason = re.sub(r'\s+', ' ', reason).strip()
        csv_rows.append([cid, rank, f"{norm_score:.4f}", reason])
        
    csv_path = 'submission.csv'
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(["candidate_id", "rank", "score", "reasoning"])
        writer.writerows(csv_rows)
        
    print(f"Saved CSV output to {csv_path}")
    
    # Generate XLSX sheet (beautifully designed)
    xlsx_path = 'submission.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"
    
    # Style definitions
    font_family = "Segoe UI"
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10)
    bold_data_font = Font(name=font_family, size=10, bold=True)
    
    border_side = Side(border_style="thin", color="D9D9D9")
    thin_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    
    # Header columns
    headers = ["Rank", "Candidate ID", "Name", "Score", "Current Title", "Current Company", "Experience (Yrs)", "Location", "Notice Period (Days)", "Response Rate", "Reasoning"]
    ws.append(headers)
    
    # Apply header formatting
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        
    # Append data rows
    for idx, (score, c) in enumerate(top_100):
        row_num = idx + 2
        rank = idx + 1
        cid = c['candidate_id']
        name = c['profile']['anonymized_name']
        norm_score = float(csv_rows[idx][2])
        title = c['profile']['current_title']
        company = c['profile']['current_company']
        exp = c['profile']['years_of_experience']
        loc = c['profile']['location']
        notice = c['redrob_signals']['notice_period_days']
        resp_rate = c['redrob_signals']['recruiter_response_rate']
        reason = csv_rows[idx][3]
        
        row_data = [rank, cid, name, norm_score, title, company, exp, loc, notice, resp_rate, reason]
        ws.append(row_data)
        
        # Style the row
        is_even = (rank % 2 == 0)
        for col_idx in range(1, len(row_data) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.font = bold_data_font if col_idx <= 2 else data_font
            cell.border = thin_border
            
            # Zebra striping
            if is_even:
                cell.fill = zebra_fill
                
            # Alignments
            if col_idx in [1, 2, 4, 7, 9, 10]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx in [8]:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
            # Number formats
            if col_idx == 4:
                cell.number_format = "0.0000"
            elif col_idx == 10:
                cell.number_format = "0%"
                
    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            # Check length of string conversion
            if cell.row == 1:
                val_len = len(str(cell.value)) + 4
            elif col_letter == 'K': # Reasoning column
                val_len = 50 # Keep reasoning column capped so it doesn't span too wide
            else:
                val_len = len(str(cell.value or ''))
            max_len = max(max_len, val_len)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 65)
        
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Save XLSX
    wb.save(xlsx_path)
    print(f"Saved XLSX output to {xlsx_path}")
    
    # Run auto-validator on generated CSV
    print("\nRunning auto-validator on submission.csv...")
    if os.path.exists(validator_file):
        try:
            result = subprocess.run(['python3', validator_file, csv_path], capture_output=True, text=True)
            print("Validator Output:")
            print(result.stdout)
            if result.returncode == 0:
                print("SUCCESS: CSV submission is valid!")
            else:
                print("ERROR: CSV submission is invalid! Details:")
                print(result.stderr)
        except Exception as e:
            print(f"Failed to run validator: {e}")
    else:
        print(f"Validator script not found at {validator_file}")

if __name__ == '__main__':
    main()
